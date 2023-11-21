import argparse
from web_scraping_scripts.filter_data import DataFilter
from web_scraping_scripts.content import WebContent
from web_scraping_scripts.workbook import Workbook
from web_scraping_scripts.analysis import analyse_data
from web_scraping_scripts.location import Location

# Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import sys
import openpyxl
import configparser

config = configparser.ConfigParser()
config.read("config.ini")

class WebScraping:
    MAX_DISPLACEMENT = None

    def __init__(self, base_url, site, db, wb, rent=None, analysis=None, market=None, filter_text=None):
        self.market = market
        self.base_url = base_url        
        # Contain the url link corressponding to each of the family mart location in location.txt file
        self.site_name = site

        self.web_scraper = []
        self.database_content = []
        
        self.check_last_link = False
        self.check_no_listing = False
        self.check_is_finished = False      # Being called 3 times: 1. There is no listing 2. There is maximum numbeer of retry reached 3. End of the page

        self.analysis = analysis
        # self.df = DataFilter(self.core_address_text, self.filter_text, self.listing_type_text)
        self.df = DataFilter(filter_text)
        self.cont = WebContent()
        self.database = db
        self.workbook = wb
        self.rental = rent

        print(f"\nVariable(base_url={base_url}, site={site}, db={db}, market={market}, filter_text={filter_text})\n")

# Get next page link and rerun
def iterate_page_propertyguru(ws, index, database):
    # print(f"INDEX {index}")
    soup = ws.cont.web_content
    family_mart_coordinates = ws.df.family_mart_coordinates[index]

    check_last = False
    get_disabled = soup.find_all(class_='pagination-next disabled')

    if get_disabled:
        print("last page reached")
        check_last = True

    # Scrape the current page information
    ws.check_no_listing = database.extract_data(soup, WebScraping.MAX_DISPLACEMENT, family_mart_coordinates)
    database.get_all()

    print("Last page? ", check_last)

    # No listing found
    if ws.check_no_listing:
        print("There is no listing available\n")
        ws.check_is_finished = True 
        return
    # Assign a new link to continue scraping
    if check_last:   # Last page
        print("There is only one page or is in last page\n")      
        ws.check_is_finished = True     # Last page
        return
    listing_pagination = soup.find(class_='pagination-next').a['href']

    # Get the next link and rerun
    second_index = listing_pagination.index('/', listing_pagination.index('/')+1)
    link = listing_pagination[second_index:]
 
    next_page_link = ws.base_url[:-2] + link
    print("Next link: ", next_page_link)
    ws.cont.connect(next_page_link)

    print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
    if ws.cont.is_retry_maximum:
        ws.check_is_finished = True
        return
    iterate_page_propertyguru(ws, index, database)

def web_scraping_propertyguru(ws):
    """
    The main function that begins the process from web-scraping, to store data in excel.

    Input: WebScraping object that contains - base url, market type, filter .txt file, family mart .txt file and listing type .txt file

    Process: 1. Store all the filter text information in DataFilter object
             2. Form a functional url based on the (name of the family mart store (without FamilyMart), listing type, and property type) and store in the RentalURLs object.
             3. For each of the url, establish it with the connection, and the loop through all the possible page it got in the website.
             4. Store the data in a .xlsx file.
             5. From the file, summarize the data by mean, median, etc.

    Output: A .xlsx file with complete list of every family mart information.

    """
    print("Extracting the filter...\n")
    ws.df.extract_all()
    print("Extract done!\n")
    print(ws.df.get_all())
  
    ws.workbook.workflows = ws.df.locations

    print("Assigning rental url...\n")
    # Initialize and store the **rental_url** objects in web_scraper list
    for index in range(ws.df.location_amount):
        rent = ws.rental(ws.base_url, ws.market)
        rent.url = (ws.df, index)

        ws.web_scraper.append(rent)
        print(f"{index+1} out of {ws.df.location_amount} being assigned.")
    print("Finish assigning.\n")

    print("Begin the scraping process\n")    
    # Each iteration if the scraping of each given link based on FM addresses
    for index, web in enumerate(ws.web_scraper):
        ws.check_is_finished = False

        print(f"This is the {index+1} location out of {len(ws.web_scraper)}")

        if index == len(ws.web_scraper)-1:
            print("This is the last link")
            ws.check_last_link = True
        # print(f"{web.__str__()}")

        ws.cont.connect(web.url)        # This store the web_content if connection is established

        # Print rental str, which is the url
        print(f"{web.__str__()}")
        database = ws.database()

        print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
        if ws.cont.is_retry_maximum:
            print("Failed to connect at first attempt.\n")
            ws.database_content.append(database) 
            print("Finished, writing to csv...\n")
            ws.workbook.store_data_propertyguru(config['File']['output_file_propertyguru'], index, ws.database_content[index])
            continue 

        iterate_page_propertyguru(ws, index, database)         # This scrape the information and get the next link for another loop
        ws.database_content.append(database)    # database_content now contains scraping information of every listing found from the link given.

        if ws.check_is_finished:
            print("Finished, writing to csv...\n")
            # print(f"content: {ws.workbook.workflows}")
            file = config['File']['output_file_propertyguru']
            # print(ws.workbook.workflows)
            ws.workbook.store_data_propertyguru(file, index, ws.database_content[index])
          
    # print("End? ", ws.check_last_link)
    # if ws.check_last_link:      # Last last last link 
    #     print(f"Data saved in {file}")


    if ws.analysis == "yes":
        analyse_data(file)

    print("Exiting...\n")
    sys.exit()

def all_property_type_in_url(ws):
    # Initialize and store the **rental_url** objects in web_scraper list
    for index in range(ws.df.location_amount):
        rent = ws.rental(ws.base_url, ws.market)
        rent.url = (ws.df, index)

        ws.web_scraper.append(rent)
        print(f"{index+1} out of {ws.df.location_amount} being assigned.")
    print("Finish assigning.\n")

def one_property_type_in_url(ws):
    temp = []
    # Initialize and store the **rental_url** objects in web_scraper list
    for index in range(ws.df.location_amount):  # location
        for types in ws.df.commercial_type_sep:       # property type 

            rent = ws.rental(ws.base_url, types, ws.df.listing_type[0])
            rent.url = (ws.df, index)
            print(rent.url)
            temp.append(rent)

            print(f"{index+1} out of {ws.df.location_amount} being assigned ({types}).")
        ws.web_scraper.append(temp)
        temp = []
    print("Finish assigning.\n")    

def iterate_page_edgeprop(ws, index, database):
    soup = ws.cont.web_content

    family_mart_coordinates = ws.df.family_mart_coordinates[index]

    check_last = False
    get_disabled = soup.find_all(class_='page-item css-q7lffx')

    if get_disabled:
        print("last page reached")
        check_last = True
    # print(soup)
    # Scrape the current page information
    ws.check_no_listing = database.extract_data(soup, WebScraping.MAX_DISPLACEMENT, ws.base_url, family_mart_coordinates)
    database.get_all()

    print("Last page? ", check_last)

    # No listing found
    # if ws.check_no_listing:
    #     print("There is no listing available\n")
    #     ws.check_is_finished = True 
    #     return
    # Assign a new link to continue scraping
    if check_last:   # Last page
        print("There is only one page or is in last page\n")      
        ws.check_is_finished = True     # Last page
        return
    
    try:
        listing_pagination = soup.find('a', class_='page-link page-link-icon')['href']
    except:
        print("There is only one page")
        return

    # Get the next link and rerun
    second_index = listing_pagination.index('/', listing_pagination.index('/')+1)
    link = listing_pagination[second_index:]
 
    next_page_link = ws.base_url[:-2] + link
    print("Next link: ", next_page_link)
    ws.cont.selenium_connect_edgeprop(next_page_link)

    print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
    if ws.cont.is_retry_maximum:
        ws.check_is_finished = True
        return
    iterate_page_edgeprop(ws, index, database)

def web_scraping_edgeprop(ws):
    file = None
    """
    The main function that begins the process from web-scraping, to store data in excel.

    Input: WebScraping object that contains - base url, market type, filter .txt file, family mart .txt file and listing type .txt file

    Process: 1. Store all the filter text information in DataFilter object
             2. Form a functional url based on the (name of the family mart store (without FamilyMart), listing type, and property type) and store in the RentalURLs object.
             3. For each of the url, establish it with the connection, and the loop through all the possible page it got in the website.
             4. Store the data in a .xlsx file.
             5. From the file, summarize the data by mean, median, etc.

    Output: A .xlsx file with complete list of every family mart information.

    """
    print("Extracting the filter...\n")
    ws.df.extract_all()
    print("Extract done!\n")
    ws.workbook.workflows = ws.df.locations

    print("Assigning rental url...\n")

    """
    If the property type cannot be in same url at the same time, ned to provide separate link.
    """
    if ws.df.commercial_type:
        all_property_type_in_url(ws)
    else:
        one_property_type_in_url(ws)

    # print(ws.web_scraper)

    print("Begin the scraping process\n")    
    # Each iteration if the scraping of each given link based on FM addresses
    database_temp = []
    for index, location in enumerate(ws.web_scraper):
        for location_url in location:
            # ws.check_is_finished = False

            print(f"This is the {index+1} location out of {len(ws.web_scraper)}")

            if index == len(ws.web_scraper)-1:
                print("This is the last link")
                ws.check_last_link = True
            # print(f"{web.__str__()}")

            ws.cont.selenium_connect_edgeprop(location_url.url)        # This store the web_content if connection is established

            # Print rental str, which is the url
            print(f"{location_url.__str__()}")

            database = ws.database()

            print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
            if ws.cont.is_retry_maximum:
                print("Failed to connect at first attempt.\n")
                database_temp.append(database) 
                # print("Finished, writing to csv...\n")
                # ws.workbook.store_data_edgeprop(config['File']['output_file_edgeprop'], index, ws.database_content[index])
                continue

            if ws.cont.web_content is None:
                print("No listing available\n")
                database_temp.append(database)
                continue
            iterate_page_edgeprop(ws, index, database)         # This scrape the information and get the next link for another loop
            database_temp.append(database)
        ws.database_content.append(database_temp)    # database_content now contains scraping information of every listing found from the link given.
        print("THIS IS: ", ws.database_content[index][0].name)
        file = config['File']['output_file_edgeprop']
        # if ws.check_is_finished:
        print(f"Finished, writing to csv named {file}...\n")
        # print(f"content: {ws.workbook.workflows}")
        print("test2", index, "and", ws.database_content[index])
        ws.workbook.store_data_edgeprop(file, index, ws.database_content[index])

        database_temp = []

    ws.workbook.store_fail(file, ws.cont.connection_fail)

    if ws.analysis == "yes":
        analyse_data(file)

    print("Exiting...\n")
    sys.exit()

def iterate_page_iproperty(ws, index, database, url, page=1):
    soup = ws.cont.web_content

    family_mart_coordinates = ws.df.family_mart_coordinates[index]

    check_last = False
    get_disabled = soup.find_all("a", attrs={'aria-label':'Go to next page'})

    if not get_disabled:
        print("last page reached")
        check_last = True

    # Scrape the current page information
    ws.check_no_listing = database.extract_data(soup, WebScraping.MAX_DISPLACEMENT, ws.base_url, family_mart_coordinates)
    database.get_all()

    print("Last page? ", check_last)

    # Assign a new link to continue scraping
    if check_last:   # Last page
        print("There is only one page or is in last page\n")      
        ws.check_is_finished = True     # Last page
        return

    page += 1
    # Get the next link and rerun
    # second_index = listing_pagination.index('/', listing_pagination.index('/')+1)
    second_index = url + '&page=' + str(page)
    link = second_index
 
    # next_page_link = ws.base_url[:-2] + link
    print("Next link: ", link)
    ws.cont.selenium_connect_iproperty(link)

    print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
    if ws.cont.is_retry_maximum:
        ws.check_is_finished = True
        return
    iterate_page_iproperty(ws, index, database, url, page)

def web_scraping_iproperty(ws):
    """
    The main function that begins the process from web-scraping, to store data in excel.

    Input: WebScraping object that contains - base url, market type, filter .txt file, family mart .txt file and listing type .txt file

    Process: 1. Store all the filter text information in DataFilter object
             2. Form a functional url based on the (name of the family mart store (without FamilyMart), listing type, and property type) and store in the RentalURLs object.
             3. For each of the url, establish it with the connection, and the loop through all the possible page it got in the website.
             4. Store the data in a .xlsx file.
             5. From the file, summarize the data by mean, median, etc.

    Output: A .xlsx file with complete list of every family mart information.

    """
    print("Extracting the filter...\n")
    ws.df.extract_all()
    print("Extract done!\n")
    ws.workbook.workflows = ws.df.locations

    print("Assigning rental url...\n")

    """
    If the property type cannot be in same url at the same time, ned to provide separate link.
    """
    if ws.df.commercial_type:
        all_property_type_in_url(ws)
    else:
        one_property_type_in_url(ws)

    # print(ws.web_scraper)

    print("Begin the scraping process\n")    
    # Each iteration if the scraping of each given link based on FM addresses
    database_temp = []
    for index, location in enumerate(ws.web_scraper):
        for location_url in location:
            # ws.check_is_finished = False

            print(f"This is the {index+1} location out of {len(ws.web_scraper)}")

            if index == len(ws.web_scraper)-1:
                print("This is the last link")
                ws.check_last_link = True
            # print(f"{web.__str__()}")

            ws.cont.selenium_connect_iproperty(location_url.url)        # This store the web_content if connection is established

            # Print rental str, which is the url
            print(f"{location_url.__str__()}")

            database = ws.database()

            print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
            if ws.cont.is_retry_maximum:
                print("Failed to connect at first attempt.\n")
                database_temp.append(database) 
                # print("Finished, writing to csv...\n")
                # ws.workbook.store_data_edgeprop(config['File']['output_file_edgeprop'], index, ws.database_content[index])
                continue

            if ws.cont.web_content is None:
                print("No listing available\n")
                database_temp.append(database)
                continue
            iterate_page_iproperty(ws, index, database, location_url.url, 1)         # This scrape the information and get the next link for another loop
            database_temp.append(database)
        ws.database_content.append(database_temp)    # database_content now contains scraping information of every listing found from the link given.
        # print("THIS IS: ", ws.database_content[index][0].name)
        file = config['File']['output_file_iproperty']
        # if ws.check_is_finished:
        print(f"Finished, writing to csv named {file}...\n")
        # print(f"content: {ws.workbook.workflows}")
        print("test2", index, "and", ws.database_content[index])


        # ws.database_content[index] -> [[a, b, c], [a, b, c], [a, b, c]]
        # the number of listings -> [1, 2, 3]
        ws.workbook.store_data_iproperty(file, index, ws.database_content[index])

        database_temp = []


    if ws.analysis == "yes":
        analyse_data(file)

    print("Exiting...\n")
    sys.exit()

# Get next page link and rerun
def iterate_page_hartamas(ws, database):
    """
    Iterate through the page of the website until the last page / connection is blocked
    """

    soup = ws.cont.web_content

    check_last = False
    get_disabled = soup.find('a', class_='rh_pagination__btn rh_pagination__next') == None
    print(get_disabled)
    if get_disabled:
        print("last page reached")
        check_last = True

    # Scrape the current page information
    ws.check_no_listing = database.extract_data(soup)
    database.get_all()

    print("Last page? ", check_last)

    # No listing found
    if ws.check_no_listing:
        print("There is no listing available\n")
        ws.check_is_finished = True 
        return
    # Assign a new link to continue scraping
    if check_last:   # Last page
        print("There is only one page or is in last page\n")      
        ws.check_is_finished = True     # Last page
        return
    listing_pagination = soup.find('a', class_="rh_pagination__btn rh_pagination__next")['href']

    print("Next link: ", listing_pagination)
    ws.cont.connect(listing_pagination)

    print("Maximum attempted reached? ", ws.cont.is_retry_maximum)
    if ws.cont.is_retry_maximum:
        ws.check_is_finished = True
        return
    iterate_page_hartamas(ws, database)

def web_scraping_hartamas(ws):
    """
    Perform web scraping in Hartamas website
    The main function that begins the process from web-scraping, to store data in excel.

    Input: WebScraping object that contains - base url, database (required) 
                                              market type, filter .txt file, family mart .txt file and listing type .txt file  (optional)

    Process: 1. Store all the filter text information in DataFilter object
             2. Form a functional url based on the (name of the family mart store (without FamilyMart), listing type, and property type) and store in the RentalURLs object. (if there is any filter)
             3. For each of the url, establish it with the connection, and the loop through all the possible page it got in the website. (in the iterate_page() function)
             4. Store the data in a .xlsx file.
             5. From the file, summarize the data by mean, median, etc.

    Output: A .xlsx file with complete list of every family mart information.
    """
    # ws.workbook = Workbook()
    print("Assigning rental url...\n")

    print("Begin the scraping process\n")    

    ws.cont.connect(ws.base_url)        # This store the web_content if connection is established
    # print(ws.cont.web_content)

    database = ws.database

    print("Maximum attempted reached? ", ws.cont.is_retry_maximum)

    if ws.cont.is_retry_maximum:
        print("Failed to connect after several attempts.\n")

        print("Finished, writing to csv...\n")
        ws.workbook.store_data_hartamas("output/hartamas.xlsx", database)

    iterate_page_hartamas(ws, database)         # This scrape the information and get the next link for another loop

    if ws.check_is_finished:
        print("Finished, writing to csv...\n")
        # print(f"content: {ws.workbook.workflows}")
        file = "output/hartamas.xlsx"
        ws.workbook.store_data_hartamas(file, database)

        print(f"Data saved in {file}")
        print("Exiting...\n")
        sys.exit()

    # analyse_data(file)


def set_constant(config):
    """
    Set up all the constant in this script
    """
    WebScraping.MAX_DISPLACEMENT = float(config['Constant']['max_displacement'])
    Location.API_KEY = config['API']['location_API']
    Location.R = float(config['Constant']['earth_radius']) 
    WebContent.RETRY_AMOUNT = int(config['Constant']['retry_attempts'])

def opt():
    parser = argparse.ArgumentParser()

    parser.add_argument("-s", "--site", type=str, default="propertyguru")
    parser.add_argument("-con", "--config", type=str, default="config.ini")
    parser.add_argument("-m", "--market", type=str, default="COMMERCIAL")
    parser.add_argument("-a", "--analysis", type=str, default="no")

    return parser.parse_args()

def main(args):
    """
    Select which site to scrap
    Input: args.site - The site to scrap
    Output: .xlsx file containing the scrap information of the site and its summary
    """
    set_constant(config)

    # print(args.filter)
    if args.site == "propertyguru":
        from web_scraping_scripts.database_propertyguru import Database
        from web_scraping_scripts.rental_url_propertyguru import RentalURLs
        wb = Workbook(('Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement'))
        ws = WebScraping(config['Link']['base_url_propertyguru'], args.site, Database, wb, RentalURLs, args.analysis, args.market, config['File']['filter_file_propertyguru'])
        web_scraping_propertyguru(ws)
    elif args.site == "hartamas":
        from web_scraping_scripts.database_hartamas import Database
        wb = Workbook(("Name", "Address", "Size", "Storey", "Psf", "Reference"))
        ws = WebScraping(config['Link']['base_url_hartamas'], args.site, Database, wb)
        web_scraping_hartamas(ws)
    elif args.site == "edgeprop":
        from web_scraping_scripts.database_edgeprop import Database
        from web_scraping_scripts.rental_url_edgeprop import RentalURLs
        wb = Workbook(('Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement'))
        ws = WebScraping(config['Link']['base_url_edgeprop'], args.site, Database, wb, RentalURLs, args.analysis, None, config['File']['filter_file_edgeprop'])
        # ws.cont.connect(ws.base_url)
        web_scraping_edgeprop(ws)
    elif args.site == "iproperty":
        from web_scraping_scripts.database_iproperty import Database
        from web_scraping_scripts.rental_url_iproperty import RentalURLs
        wb = Workbook(('Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement'))
        ws = WebScraping(config['Link']['base_url_iproperty'], args.site, Database, wb, RentalURLs, args.analysis, None, config['File']['filter_file_iproperty'])
        # ws.cont.connect(ws.base_url)
        web_scraping_iproperty(ws)      
    
    else:
        print(f"There site {args.site} is not supported")
        sys.exit(0)

if __name__ == "__main__":
    args = opt()
    print(args)

    main(args)

    # analyse_data("sheets/Family Mart Rental Information2.xlsx")