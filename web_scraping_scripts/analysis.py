import argparse
import openpyxl
import copy
import configparser
import numpy as np

config = configparser.ConfigParser()
config.read("config.ini")


def analyse_data(file):
    """
    Calculate the data from the result in .xlsx file
    Input: Output file
    Output: Statistic summary such as mean, median, standard deviation, etc.
    """

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    store_name_column = 0
    price_column = 3
    size_column = 4
    psf_column = 5

    store_name_values = []
    price_values = []
    size_values = []
    psf_values = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        store_name_value = row[store_name_column]
        price_value = row[price_column]
        size_value = row[size_column]
        psf_value = row[psf_column]

        store_name_values.append(store_name_value)
        price_values.append(price_value)
        size_values.append(size_value)
        psf_values.append(psf_value)

    important_value_dict = {}
    data = {}

    value_temp = []

    checker = config['Constant']['family_mart_checker']
    record_data = 0

    store_name = None
    # important_value_dict = OrderedDict()
    for index, name in enumerate(store_name_values):
        # print(f"data here is {data}")
        # print(name)
        if checker in str(name):     # run when the name contains FamilyMart
            store_name = name
            # continue for the first time 
            if record_data >= 1:        # for more than first time, record data
                important_value_dict["name"] = name
                value_temp.append((price_values[index], size_values[index], psf_values[index]))
                        
            record_data += 1
        else:
            if record_data >= 1:
                # print(f"Value is {value_temp}")

                data[store_name] = copy.copy(value_temp)
                # print(f"data is {data}")
                value_temp.clear()
                record_data = 0
            record_data = 0 # means it finished loop the previous listings and ned to giv new one

    statistic = {}
 
    """
    data = 
    {
    "Family Mart Store Name": [],
    "Family Mart Store Name": [(price, size, psf)], 
    "Family Mart Store Name": [(price, size, psf),(price, size, psf)...] ... 
    }
    """
    number_of_parameter = 16
    for key, value in data.items():
        if key == "FamilyMart Mid Valley":
            print(value)
        price = []
        size = []
        psf = []
        if not value:
            price_mean = None
            size_mean = None
            psf_mean = None
            unavailable = np.zeros(number_of_parameter)
            statistic[key] = unavailable
        else:
            for index in range(len(value)):
                p, s, ps = value[index]
                if key == "FamilyMart Mid Valley":
                    print(f"price is {p}, size is {s}, psf is {ps}")

                if None in (p, s, ps):
                    continue
                else:
                    # print("here")
                    price.append(p)
                    size.append(s)
                    psf.append(ps)


                prices = np.array(price)
                sizes = np.array(size)
                psfs = np.array(psf)
                
                # Amount
                amount = len(price)
                
                # Mean
                price_mean = np.mean(prices).round(1)
                size_mean = np.mean(sizes).round(1)
                psf_mean = np.mean(psfs).round(1)

                # Median
                price_median = np.median(prices)
                size_median = np.median(sizes)
                psf_median = np.median(psfs)

                # 0.25 Percentile
                price_25percentile = np.quantile(prices, 0.25)
                size_25percentile = np.quantile(sizes, 0.25)
                psf_25percentile = np.quantile(psfs, 0.25)

                # 0.75 Percentile
                price_75percentile = np.quantile(prices, 0.75)
                size_75percentile = np.quantile(sizes, 0.75)
                psf_75percentile = np.quantile(psfs, 0.75)

                # Standard deviation
                price_std = np.std(prices).round(2)
                size_std = np.std(sizes).round(2)
                psf_std = np.std(psfs).round(2)


            statistic[key] = np.array([price_mean, price_median, price_25percentile, price_75percentile, price_std, 
                            size_mean, size_median, size_25percentile, size_75percentile, size_std,
                            psf_mean, psf_median, psf_25percentile, psf_75percentile, psf_std, amount])
            price_mean = price_median = price_25percentile = price_75percentile = price_std = size_mean = size_median = size_25percentile = size_75percentile = size_std = psf_mean = psf_median = psf_25percentile = psf_75percentile = psf_std = amount = 0
            # print(f"{key} and {prices}")
                               
    print(f"statistic is {statistic}")

    sheet = workbook.create_sheet(title="Summary")

    head = ('Price', 'Size', 'Psf')
    header = ('Mean', 'Median', '.25 Percentile', '.75 Percentile', 'Standard Deviation')

    for i in range(1, len(head)+1):
        sheet.cell(row=1, column=i*5-3, value=str(head[i-1]))

    sheet.cell(row=2, column=1, value='FamilyMart Store')
    for i in range(1, len(header)*3+1):
        sheet.cell(row=2, column=i+1, value=str(header[(i-1)%5]))
    sheet.cell(row=2, column=17, value='Number of Listing')

    for key, value in statistic.items():
        # print(value.size)
        sheet.append([key, value[0], value[1], value[2], value[3], value[4], value[5], value[6], value[7], value[8], value[9], value[10], value[11], value[12], value[13], value[14], value[15]])

    workbook.save(file)
    workbook.close()

def opt():
    parser = argparse.ArgumentParser()

    parser.add_argument("-f", "--file", type=str)

    return parser.parse_args()

if __name__ == '__main__':
    args = opt()
    analyse_data(args.file)