import openpyxl
import os

class Workbook:
    def __init__(self, headers, workflows=None):
        self._headers = headers
        self._workflows = workflows

    @property
    def headers(self):
        return self._headers

    @headers.setter
    def headers(self, headers):
        self._headers = headers

    @property
    def workflows(self, index=None):
        return self._workflows if index is None else self._workflows[index]

    @workflows.setter
    def workflows(self, workflows):
        self._workflows = workflows

    def create_multiple_sheets(self, workbook_name, worksheet_new_name):
        # Open or load a workbook if existed
        if os.path.exists(workbook_name):
            workbook = openpyxl.load_workbook(workbook_name)
        else:
            workbook = openpyxl.Workbook()

        if workbook.active.title == "Sheet1":
            print("Only have sheet1 active")
            sheet = workbook.active
            sheet.title = worksheet_new_name
        # Create or use active worksheet
        elif workbook.active.title != worksheet_new_name:
            # print(f"{workbook.active.title} and {worksheet_new_name} is {workbook.active.title != worksheet_new_name}")
            sheet = workbook.create_sheet(title=worksheet_new_name)
        else:           # Might run into data loss
            print("same name, replacing")
            sheet = workbook.active
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
        
                # Add header
        for i, header in enumerate(self.headers, 1):
            sheet.cell(row=1, column=i, value=str(header).capitalize())

        return workbook, sheet

    # def open_current_sheet_propertyguru(self, workbook_name, worksheet_new_name, index):
    #     # Open or load a workbook if existed
    #     if os.path.exists(workbook_name):
    #         workbook = openpyxl.load_workbook(workbook_name)
    #     else:
    #         workbook = openpyxl.Workbook()

    #     sheet = workbook.active
    #     # self.headers = ('Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement')
    #     for i, header in enumerate(self.headers, 1):
    #         sheet.cell(row=1, column=i, value=str(header).capitalize())
    #     sheet.append([(index+1)])
    #     sheet.append([worksheet_new_name])

    #     return workbook, sheet
    
    # def open_current_sheet_hartamas(self, workbook_name):
    #     # Open or load a workbook if existed
    #     if os.path.exists(workbook_name):
    #         workbook = openpyxl.load_workbook(workbook_name)
    #     else:
    #         workbook = openpyxl.Workbook()

    #     sheet = workbook.active
    #     # self.headers = ("Name", "Address", "Size", "Storey", "Psf", "Reference")
    #     for i, header in enumerate(self.headers, 1):
    #         sheet.cell(row=1, column=i, value=str(header).capitalize())

    #     return workbook, sheet

    def open_current_sheet(self, workbook_name, worksheet_new_name=None, index=None):
        # Open or load a workbook if existed
        if os.path.exists(workbook_name):
            workbook = openpyxl.load_workbook(workbook_name)
        else:
            workbook = openpyxl.Workbook()

        sheet = workbook.active
        # self.headers = ('Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement')
        for i, header in enumerate(self.headers, 1):
            sheet.cell(row=1, column=i, value=str(header).capitalize())

        if index is not None:
            sheet.append([(index+1)])
            sheet.append([worksheet_new_name])

        return workbook, sheet    

    """
    openpyxl
    """
    def store_data_propertyguru(self, workbook_name, index, data):
        listing = len(data.name)
        print(listing)
        print(f"workflow is {self.workflows[index]}\n")
        worksheet_new_name = self.workflows[index] 
        # print("Whole is: " + worksheet_new_name)

        # workbook, sheet = self.create_multiple_sheets(workbook_name, worksheet_new_name)
        workbook, sheet = self.open_current_sheet(workbook_name, worksheet_new_name, index)

        # Add data
        data.get_all()

        for i in range(listing):
            try:
                price = float(data.price[i])
            except:
                price = None

            try:
                size = float(data.size[i])
            except:
                size = None

            try:
                psf = float(data.psf[i])
            except:
                psf = None

            try:
                displacement = float(data.displacement[i])
            except:
                displacement = None

            sheet.append([str(self.workflows[index]), str(data.name[i]), str(data.description[i]).strip("['']"), price
            , size, psf, str(data.reference[i]), str(data.address[i]), displacement])

        sheet.append(["Number of listing:", listing])
        sheet.append([" "])
        # sheet.cell(row=sheet.max_row+1, column='A', value=str(ds.get_index()))
        workbook.save(workbook_name)
        # Close workbook
        workbook.close()
        print(f"Successfully writed the listing for {self.workflows[index]}\n")

    def store_data_hartamas(self, workbook_name, data):
        listing = len(data.name)

        # workbook, sheet = self.create_multiple_sheets(workbook_name, worksheet_new_name)
        workbook, sheet = self.open_current_sheet(workbook_name)

        # Print data
        data.get_all()

        for i in range(listing):
            try:
                size = str(data.size[i])
            except:
                size = None

            try:
                psf = str(data.psf[i])
            except:
                psf = None

            try:
                name = str(data.name[i])
            except:
                name = None
            
            try:
                address = str(data.address[i])
            except:
                address = None
            
            try:
                storey = str(data.storey[i])
            except:
                storey = None
            
            try:
                reference = str(data.reference[i])
            except:
                reference = None

            # if i in list_to_filter:
            #     continue
            # print(i)
            sheet.append([name, address, size, storey, psf, reference])

        #     # index += 1
        sheet.append(["Number of listing:", listing])
        # sheet.append([" "])
        # sheet.cell(row=sheet.max_row+1, column='A', value=str(ds.get_index()))
        workbook.save(workbook_name)
        # Close workbook
        workbook.close()
        print(f"Successfully writed the listing\n")

    def store_data_edgeprop(self, workbook_name, index, data):
        print("test", data)

        print(f"workflow is {self.workflows[index]}\n")
        worksheet_new_name = self.workflows[index] 
        # print("Whole is: " + worksheet_new_name)

        # workbook, sheet = self.create_multiple_sheets(workbook_name, worksheet_new_name)
        workbook, sheet = self.open_current_sheet(workbook_name, worksheet_new_name, index)

        length = 0
        num_of_listings = 0
        for ind in range(len(data)):
            print("name: ", data[ind].name)
            listing = len(data[ind].name)
            length += listing

            # Add data
            data[ind].get_all()
            num_of_listings += data[ind].num_of_listings
            for i in range(listing):
                try:
                    price = float(data[ind].price[i])
                except:
                    price = None

                try:
                    size = float(data[ind].size[i])
                except:
                    size = None

                try:
                    psf = float(data[ind].psf[i])
                except:
                    psf = None

                try:
                    displacement = float(data[ind].displacement[i])
                except:
                    displacement = None

                sheet.append([str(self.workflows[index]), str(data[ind].name[i]), str(data[ind].description[i]).strip("['']"), price
                , size, psf, str(data[ind].reference[i]), str(data[ind].address[i]), displacement])

        sheet.append([f"Number of listing: {length}", f"Number of all listings: {num_of_listings}"])
        sheet.append([" "])
        # sheet.cell(row=sheet.max_row+1, column='A', value=str(ds.get_index()))
        workbook.save(workbook_name)
        # Close workbook
        workbook.close()
        print(f"Successfully writed the listing for {self.workflows[index]}\n")

    def store_data_iproperty(self, workbook_name, index, data):
        print("test", data)

        print(f"workflow is {self.workflows[index]}\n")
        worksheet_new_name = self.workflows[index] 
        # print("Whole is: " + worksheet_new_name)

        # workbook, sheet = self.create_multiple_sheets(workbook_name, worksheet_new_name)
        workbook, sheet = self.open_current_sheet(workbook_name, worksheet_new_name, index)

        length = 0
        num_of_listings = 0
        # ws.database_content[index] -> [[a, b, c], [a, b, c], [a, b, c]]
        # the number of listings -> [1, 2, 3]
        for ind in range(len(data)):
            print("name: ", data[ind].name)
            listing = len(data[ind].name)
            length += listing
            # num_of_listings += data[ind].num_of_listings

            # Add data
            data[ind].get_all()
            num_of_listings += data[ind].num_of_listings
            for i in range(listing):
                
                try:
                    price = float(data[ind].price[i])
                except:
                    price = None

                try:
                    size = float(data[ind].size[i])
                except:
                    size = None

                try:
                    psf = float(data[ind].psf[i])
                except:
                    psf = None

                try:
                    displacement = float(data[ind].displacement[i])
                except:
                    displacement = None

                sheet.append([str(self.workflows[index]), str(data[ind].name[i]), str(data[ind].description[i]).strip("['']"), price
                , size, psf, str(data[ind].reference[i]), str(data[ind].address[i]), displacement])

        sheet.append([f"Number of listing: {length}", f"Number of all listings: {num_of_listings}"])
        sheet.append([" "])
        # sheet.cell(row=sheet.max_row+1, column='A', value=str(ds.get_index()))
        workbook.save(workbook_name)
        # Close workbook
        workbook.close()
        print(f"Successfully writed the listing for {self.workflows[index]}\n")

    def store_fail(self, workbook, fails):
        workbook, sheet = self.open_current_sheet(workbook)
        sheet.append("Link that failed:")
        for fail in fails:
            sheet.append([fail])



       

if __name__ == '__main':
    workbook = Workbook()