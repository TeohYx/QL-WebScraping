import argparse
import openpyxl
import os
from web_scraping_scripts.analysis import analyse_data

def combine_listing(lists):
    # print(lists)
    fm_lists = {}
    for list in lists:
        if os.path.exists(list):
            workbook = openpyxl.load_workbook(list)
        else:
            raise "error loading workbook"

        workbook.active = workbook['Sheet1']
        sheet = workbook.active
        # print(sheet)

        column_obj = []

        print(f"Current workbook is {list}")
        # Run through all the file 
        # print(sheet.max_row)
        for i in range(2, sheet.max_row+1):
            first_cell = (sheet.cell(row=i, column=1).value, sheet.cell(row = i, column = 2).value)
            check_cell = sheet.cell(row = i, column = sheet.max_column)

            try:
                previous_cell = sheet.cell(row = i-1, column = sheet.max_column)
            except:
                print("previous value not available")
                previous_cell = None
            
            try:
                after_cell = sheet.cell(row = i+1, column = sheet.max_column)
                blank_cell = sheet.cell(row = i+2, column = sheet.max_column)
                number_cell = (sheet.cell(row=i, column=1).value, sheet.cell(row = i+1, column = 2).value)
            except:
                print("bound error")
                after_cell = None
                blank_cell = None
                number_cell = None
            # print(blank_cell.value)
            # If FamilyMart and 9th column is not None 
            if after_cell.value is not None and check_cell.value is None or blank_cell.value is None and check_cell.value is None:       # Indicate at least one list
                # print(after_cell.value)
                is_family_mart = True
                if first_cell in fm_lists.keys():         # To eliminate duplicates
                    continue
                fm_lists[first_cell] = []
            # If FamilyMart and 9th column is None
            # Three of the conditions where the data needs to be stored.
            elif (after_cell.value is not None and check_cell.value is not None 
                  or previous_cell.value is not None and check_cell.value is not None 
                  or previous_cell.value is None and check_cell.value is not None and after_cell.value is None):
                # Get all the data of the column 
                for j in range(1, sheet.max_column+1):   
                    cell_obj = sheet.cell(row = i, column = j)
                    column_obj.append(cell_obj.value)
                fm_lists[first_cell].append(column_obj.copy())
                column_obj.clear()
            else:
                continue
    # print(fm_lists)
    return fm_lists

def store_results(fm_lists, new_file):
    index = 1

    if os.path.exists(new_file):
        workbook = openpyxl.load_workbook(new_file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    headers = ('Store No', 'Store Name', 'Name', 'Description', 'Price', 'Size', 'Psf', 'Reference', 'Address', 'Displacement')
    for i, header in enumerate(headers, 1):
        sheet.cell(row=1, column=i, value=str(header).capitalize())

    for key, values in fm_lists.items():
        sheet.append([(index)])
        sheet.append([key[0], key[1]])
        
        for value in values:
            sheet.append(value)

        sheet.append([f"Number of listings: {len(values)}"])
        sheet.append([" "])
        index += 1

    workbook.save(new_file)
    workbook.close()
    print(f"Successfully writed the listing \n")

def opt():
    parser = argparse.ArgumentParser()

    parser.add_argument('-l', '--list', nargs='+', required=True)
    parser.add_argument('-o', '--output', type=str, required=True, help="output in xlsx. format")

    return parser.parse_args()

def main(args):
    new_file = args.output

    fm_lists = combine_listing(args.list)
    store_results(fm_lists, new_file)

    analyse_data(new_file)

if __name__ == '__main__':
    args = opt()
    main(args)

