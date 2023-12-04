import argparse
import openpyxl
import copy
import configparser
import numpy as np

config = configparser.ConfigParser()
config.read("config.ini")


def analyse_data(file):
    """
    Calculate the data from the result in .xlsx file
    Input: Output file
    Output: Statistic summary such as mean, median, standard deviation, etc.
    """

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    store_id_column = 0
    store_name_column = 1
    link_column = 7
    price_column = 4
    size_column = 5
    psf_column = 6
    site_column = 10

    store_name_values = []
    link_values = []
    price_values = []
    size_values = []
    psf_values = []
    # ALl data of site
    site_values = []
    # Unique data of site
    site_type = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        store_name_value = (row[store_id_column], row[store_name_column])
        price_value = row[price_column]
        size_value = row[size_column]
        psf_value = row[psf_column]
        link_value = row[link_column]
        site_value = row[site_column]

        store_name_values.append(store_name_value)
        price_values.append(price_value)
        size_values.append(size_value)
        psf_values.append(psf_value)
        link_values.append(link_value)
        """
        site_values holds all the site from the excel
        site_type gets the unique type of the site
        """
        site_values.append(site_value)

        if (site_type is None or site_value not in site_type) and site_value is not None:
            site_type.append(site_value)

    # for i in range(len(site_values)):
    #     if store_name_values[i][0] == 9999:
    #         print(site_values[i])

    important_value_dict = {}
    data = {}

    value_temp = []
    site_counter = []
    for i in range(len(site_type)):
        site_counter.append([site_values[i], 0])
    checker = config['Constant']['family_mart_checker']
    record_data = 0

    store_name = None
    # print(store_name_values, price_values, size_values, psf_values)
    # important_value_dict = OrderedDict()
    name_id = None
    for index, name in enumerate(store_name_values):
        # When name_id is nothing
        if name_id is None:
            name_id = name
            
        # When in same store
        if name_id == name:

            # print(f"price: {price_values[index]}, size: {size_values[index]}, psf: {psf_values[index]}")
            value_temp.append((price_values[index], size_values[index], psf_values[index]))            
            site = site_values[index]
            # print(site_type)
            for i in range(len(site_type)):
                if site == site_type[i]:
                    site_counter[i][1] += 1
        if name_id != name or index == len(store_name_values)-1:
            # if index < 86:
            #     print(index)
            #     print(name)
            #     print((copy.copy(value_temp), copy.deepcopy(site_counter)))
            # print(site_counter)
            data[name_id] = (copy.copy(value_temp), copy.deepcopy(site_counter))

            value_temp.clear()

            for i in site_counter:
                i[1] = 0
            record_data = 0   
            name_id = name       

            value_temp.append((price_values[index], size_values[index], psf_values[index]))       
            site = site_values[index]         
            for i in range(len(site_type)):
                if site == site_type[i]:
                    site_counter[i][1] += 1

    statistic = {}
    # for k, v in data.items():
    #     print(k, v[1])
    """
    data = 
    {
    "Family Mart Store Name": [],
    "Family Mart Store Name": [(price, size, psf)], 
    "Family Mart Store Name": [(price, size, psf),(price, size, psf)...] ... 
    }
    """
    number_of_parameter = 17
    total_site = []
    total_site_str = ""
    for key, value in data.items():
        # print(value)
        # if key == "FamilyMart Mid Valley":
        #     print(value)
        price = []
        size = []
        psf = []
        site = ""
        site_title = ""



        if not value:
            price_mean = None
            size_mean = None
            psf_mean = None
            unavailable = np.zeros(number_of_parameter)
            statistic[key] = unavailable
        else:
            for index in range(len(value[0])):
                p, s, ps = value[0][index]
                # if key == "FamilyMart Mid Valley":
                #     print(f"price is {p}, size is {s}, psf is {ps}")

                if None in (p, s, ps):
                    continue
                else:
                    # print("here")
                    price.append(p)
                    size.append(s)
                    psf.append(ps)


                prices = np.array(price)
                sizes = np.array(size)
                psfs = np.array(psf)
                
                # Amount
                amount = len(price)
                
                # Mean
                price_mean = np.mean(prices).round(1)
                size_mean = np.mean(sizes).round(1)
                psf_mean = np.mean(psfs).round(1)

                # Median
                price_median = np.median(prices)
                size_median = np.median(sizes)
                psf_median = np.median(psfs)

                # 0.25 Percentile
                price_25percentile = np.quantile(prices, 0.25)
                size_25percentile = np.quantile(sizes, 0.25)
                psf_25percentile = np.quantile(psfs, 0.25)

                # 0.75 Percentile
                price_75percentile = np.quantile(prices, 0.75)
                size_75percentile = np.quantile(sizes, 0.75)
                psf_75percentile = np.quantile(psfs, 0.75)

                # Standard deviation
                price_std = np.std(prices).round(2)
                size_std = np.std(sizes).round(2)
                psf_std = np.std(psfs).round(2)      

            if not total_site: 
                for i in range(len(value[1])):
                    total_site.append(0)

            for i in range(len(value[1])):
                site += (str(value[1][i][1]) + ", ")
                total_site[i] += value[1][i][1]
            site = site[:-2]
            print(total_site)
            statistic[key] = np.array([price_mean, price_median, price_25percentile, price_75percentile, price_std, 
                            size_mean, size_median, size_25percentile, size_75percentile, size_std,
                            psf_mean, psf_median, psf_25percentile, psf_75percentile, psf_std, amount, site])
            price_mean = price_median = price_25percentile = price_75percentile = price_std = size_mean = size_median = size_25percentile = size_75percentile = size_std = psf_mean = psf_median = psf_25percentile = psf_75percentile = psf_std = amount = 0
            site = ""
            # print(f"{key} and {prices}")

    # print(f"statistic is {statistic}")
    for i in site_type:
        site_title += (i + ", ")
    for i in total_site:
        total_site_str += (str(i) + ", ")
    total_site_str = total_site_str[:-2]

    sheet = workbook.create_sheet(title="Summary")

    head = ('Price', 'Size', 'Psf')
    header = ('Mean', 'Median', '.25 Percentile', '.75 Percentile', 'Standard Deviation')

    for i in range(1, len(head)+1):
        sheet.cell(row=1, column=i*5-3+1, value=str(head[i-1]))
    sheet.cell(row=2, column=1, value='FamilyMart No')
    sheet.cell(row=2, column=2, value='FamilyMart Store')
    for i in range(1, len(header)*3+1):
        sheet.cell(row=2, column=i+1+1, value=str(header[(i-1)%5]))
    sheet.cell(row=2, column=18, value='Number of Listing')
    sheet.cell(row=2, column=19, value= site_title)

    for key, value in statistic.items():
        # print(value.size)
        sheet.append([key[0], key[1], float(value[0]), float(value[1]), float(value[2]), float(value[3]), float(value[4]), float(value[5]), float(value[6]), float(value[7]), float(value[8]), float(value[9]), float(value[10]), float(value[11]), float(value[12]), float(value[13]), float(value[14]), int(value[15]), value[16]])
    sheet.cell(row=(sheet.max_row + 1), column=19, value=total_site_str)
    print("Finished summarised")
    workbook.save(file)
    workbook.close()

def opt():
    parser = argparse.ArgumentParser()

    parser.add_argument("-f", "--file", type=str)

    return parser.parse_args()

if __name__ == '__main__':
    args = opt()
    analyse_data(args.file)